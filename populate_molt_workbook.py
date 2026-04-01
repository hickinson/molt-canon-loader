#!/usr/bin/env python3
"""Production loader for Molt canon workbook full imports."""

import argparse
import datetime as dt
import hashlib
import json
import re
from collections import Counter
from copy import copy
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook

API_URL = "https://molt.church/api/canon"
RAW_IMPORT_SHEET = "Raw Import"
CONTROL_SHEET = "Controls"
START_ROW = 5
SOURCE_REGISTER_SHEET = "Source Register"
SHORTLIST_SHEET = "Shortlist Model"
ANTHOLOGY_SHEET = "Anthology Candidates"
DASHBOARD_SHEET = "Dashboard"
RAW_COLUMNS = 9
SUPPORTED_TYPES = {"prophecy", "psalm", "parable", "teaching", "liturgy", "prayer", "vision", "proverb"}


class CanonLoadError(RuntimeError):
    """Raised when a live canon load cannot be completed."""


def utc_now() -> dt.datetime:
    return dt.datetime.now(dt.timezone.utc).replace(microsecond=0)


def to_excel_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value)


def normalize_canonized_at(value: Any, anomalies: list[str], row_id: int) -> str:
    raw = to_excel_cell(value)
    if not raw:
        anomalies.append(f"row {row_id}: missing canonized_at")
        return ""

    # Accept unmodified source values but flag malformed dates.
    try:
        dt.datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except ValueError:
        anomalies.append(f"row {row_id}: malformed canonized_at '{raw}'")
    return raw


def fetch_payload(source_url: str, timeout_seconds: int) -> dict[str, Any]:
    response = requests.get(source_url, timeout=timeout_seconds)
    response.raise_for_status()
    payload = response.json()
    if not isinstance(payload, dict):
        raise CanonLoadError("API payload root is not a JSON object")
    return payload


def flatten_records(payload: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str], dict[str, int]]:
    anomalies: list[str] = []
    book = payload.get("the_great_book")
    if book is None:
        raise CanonLoadError("Payload missing required 'the_great_book' collection")
    if not isinstance(book, list):
        raise CanonLoadError("Payload field 'the_great_book' is not a list")

    rows: list[dict[str, Any]] = []
    scripture_types = Counter()
    content_seen: Counter[str] = Counter()

    for idx, item in enumerate(book, start=1):
        if not isinstance(item, dict):
            anomalies.append(f"row {idx}: non-object entry skipped (type={type(item).__name__})")
            continue

        prophet_name = to_excel_cell(item.get("prophet_name"))
        scripture_type = to_excel_cell(item.get("scripture_type"))
        content = to_excel_cell(item.get("content"))
        canonized_at = normalize_canonized_at(item.get("canonized_at"), anomalies, idx)

        if not prophet_name:
            anomalies.append(f"row {idx}: missing prophet_name")
        if not scripture_type:
            anomalies.append(f"row {idx}: missing scripture_type")
            scripture_types["<missing>"] += 1
        else:
            scripture_types[scripture_type] += 1
        if not content:
            anomalies.append(f"row {idx}: missing content")

        content_seen[content] += 1
        rows.append(
            {
                "Raw_Row_ID": len(rows) + 1,
                "Raw_Prophet_Name": prophet_name,
                "Raw_Scripture_Type": scripture_type,
                "Raw_Content": content,
                "Raw_Canonized_At": canonized_at,
            }
        )

    duplicate_non_empty = sum(1 for k, count in content_seen.items() if k and count > 1)
    if duplicate_non_empty:
        anomalies.append(f"detected {duplicate_non_empty} duplicate Raw_Content values")

    stats = {
        "source_entries": len(book),
        "imported_entries": len(rows),
        "skipped_entries": len(book) - len(rows),
    }
    return rows, anomalies, dict(scripture_types)


def clear_data_rows(ws, total_columns: int, keep_template_row: int = START_ROW) -> None:
    max_row = ws.max_row
    for row in range(keep_template_row, max_row + 1):
        for col in range(1, total_columns + 1):
            cell = ws.cell(row=row, column=col)
            if row == keep_template_row:
                # Retain style template on first data row; reset value only.
                cell.value = None
            else:
                cell.value = None


def copy_row_style(ws, from_row: int, to_row: int, total_columns: int) -> None:
    for col in range(1, total_columns + 1):
        src = ws.cell(row=from_row, column=col)
        dst = ws.cell(row=to_row, column=col)
        dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)


def write_rows(
    ws,
    rows: list[dict[str, Any]],
    source_url: str,
    snapshot_id: str,
    imported_at: str,
    import_batch_name: str,
) -> None:
    clear_data_rows(ws, total_columns=RAW_COLUMNS)

    for offset, item in enumerate(rows):
        row_num = START_ROW + offset
        if row_num > START_ROW:
            copy_row_style(ws, START_ROW, row_num, total_columns=RAW_COLUMNS)

        ws.cell(row=row_num, column=1, value=item["Raw_Row_ID"])
        ws.cell(row=row_num, column=2, value=item["Raw_Prophet_Name"])
        ws.cell(row=row_num, column=3, value=item["Raw_Scripture_Type"])
        ws.cell(row=row_num, column=4, value=item["Raw_Content"])
        ws.cell(row=row_num, column=5, value=item["Raw_Canonized_At"])
        ws.cell(row=row_num, column=6, value=source_url)
        ws.cell(row=row_num, column=7, value=snapshot_id)
        ws.cell(row=row_num, column=8, value=imported_at)
        ws.cell(row=row_num, column=9, value=import_batch_name)


def normalize_text_for_flags(content: str) -> str:
    return re.sub(r"\s+", " ", content).strip().lower()


def parse_iso_date(raw_value: str) -> tuple[dt.datetime | None, bool]:
    if not raw_value:
        return None, True
    try:
        parsed = dt.datetime.fromisoformat(raw_value.replace("Z", "+00:00"))
        return parsed.replace(tzinfo=None), False
    except ValueError:
        return None, True


def build_source_register_rows(rows: list[dict[str, Any]], anomalies: list[str]) -> list[list[Any]]:
    output: list[list[Any]] = []
    hash_counts = Counter()
    prepared = []

    for item in rows:
        content = item["Raw_Content"]
        canonized_at = item["Raw_Canonized_At"]
        normalized_type_raw = (item["Raw_Scripture_Type"] or "").strip().lower()
        normalized_type = normalized_type_raw.title() if normalized_type_raw else ""
        parsed_date, malformed_date = parse_iso_date(canonized_at)
        content_normalized = normalize_text_for_flags(content)
        digest = hashlib.md5(content_normalized.encode("utf-8")).hexdigest() if content_normalized else ""
        hash_counts[digest] += 1
        prepared.append((item, content, parsed_date, malformed_date, normalized_type_raw, normalized_type, digest))

    symbolic_tokens = {"soul", "sacred", "divine", "ritual", "memory", "world", "threshold", "light", "witness"}

    for idx, (item, content, parsed_date, malformed_date, normalized_type_raw, normalized_type, digest) in enumerate(prepared, start=1):
        preview = re.sub(r"\s+", " ", content).strip()[:160]
        words = len([w for w in re.split(r"\s+", content.strip()) if w]) if content else 0
        chars = len(content) if content else 0
        empty_or_broken = words == 0
        very_short = words < 8 and words > 0
        excessive = chars > 3000
        unsupported_type = bool(normalized_type_raw) and normalized_type_raw not in SUPPORTED_TYPES
        if unsupported_type:
            anomalies.append(f"row {idx}: unexpected scripture_type '{item['Raw_Scripture_Type']}'")

        duplicate_text = bool(digest and hash_counts[digest] > 1)
        low_substance = "Candidate" if (very_short or words < 18) else "Clear"
        token_promo = "Candidate" if "http://" in content.lower() or "https://" in content.lower() else "Clear"
        likely_quote = "Candidate" if '"' in content and words < 25 else "Clear"
        strong_symbolic = "Yes" if any(token in content.lower() for token in symbolic_tokens) else "No"
        outsider_friendly = "Yes" if words <= 80 else "No"
        high_distinctive = "Yes" if chars >= 180 else "No"
        historical = "Yes" if re.search(r"\b(19|20)\d{2}\b", content) else "No"
        native_energy = "High" if words >= 120 else "Medium" if words >= 50 else "Low"

        if empty_or_broken:
            screening_status = "Probable exclude"
            review_priority = "High"
            screening_notes = "Missing content"
        elif malformed_date or unsupported_type:
            screening_status = "Review before scoring"
            review_priority = "High"
            screening_notes = "Schema or date anomaly"
        elif very_short:
            screening_status = "Review before scoring"
            review_priority = "Medium"
            screening_notes = "Very short content"
        else:
            screening_status = "Pass to scoring"
            review_priority = "High" if strong_symbolic == "Yes" else "Medium"
            screening_notes = "Auto-pass baseline checks"

        final_status = screening_status
        row = [
            idx,  # Working_Row_ID
            item["Raw_Row_ID"],
            item["Raw_Prophet_Name"],
            item["Raw_Scripture_Type"],
            content,
            canonized_at,
            preview,
            parsed_date,
            parsed_date.year if parsed_date else None,
            parsed_date.month if parsed_date else None,
            words,
            chars,
            "Short" if words < 40 else "Medium" if words < 140 else "Long",
            normalized_type,
            digest,
            "Yes" if duplicate_text else "No",
            "Yes" if empty_or_broken else "No",
            "Yes" if very_short else "No",
            "Yes" if malformed_date else "No",
            "Yes" if unsupported_type else "No",
            "Yes" if excessive else "No",
            low_substance,
            token_promo,
            likely_quote,
            strong_symbolic,
            outsider_friendly,
            high_distinctive,
            historical,
            native_energy,
            screening_status,
            review_priority,
            screening_notes,
            None,  # Manual override
            final_status,
            "Auto Loader",
            utc_now().replace(tzinfo=None),
            1,
            None,
        ]
        output.append(row)
    return output


def score_row(words: int, chars: int, symbolic: str, distinctive: str, historical: str, outsider: str) -> tuple[int, int, int, int, int, int]:
    literary = max(1, min(5, 1 + chars // 220))
    symbolic_score = 4 if symbolic == "Yes" else 2
    theological = 3 if symbolic == "Yes" else 2
    distinct = 4 if distinctive == "Yes" else 2
    outsider_score = 4 if outsider == "Yes" else 2
    history_score = 4 if historical == "Yes" else 2
    return literary, symbolic_score, theological, distinct, outsider_score, history_score


def build_shortlist_rows(source_rows: list[list[Any]]) -> list[list[Any]]:
    shortlist = []
    for row in source_rows:
        final_status = row[33]
        if final_status != "Pass to scoring":
            continue
        words, chars = row[10], row[11]
        literary, symbolic_score, theological, distinct, outsider_score, history_score = score_row(
            words, chars, row[24], row[26], row[27], row[25]
        )
        total = literary + symbolic_score + theological + distinct + outsider_score + history_score
        shortlist.append(
            [
                row[0],  # Working_Row_ID
                row[2],  # prophet
                row[13],  # normalized type
                row[7],  # parsed date
                row[6],  # preview
                final_status,
                row[30],  # review priority
                literary,
                symbolic_score,
                theological,
                distinct,
                outsider_score,
                history_score,
                total,
                row[28],  # native energy
                "Witness",
                "Entry",
                "Anthology Candidate" if total >= 20 else "Review",
                "Threshold",
                None,
                "Yes" if total >= 20 else "No",
                "Auto-scored by loader; editorial review required.",
            ]
        )
    return shortlist


def build_anthology_rows(shortlist_rows: list[list[Any]]) -> list[list[Any]]:
    candidates = [r for r in shortlist_rows if r[20] == "Yes"]
    candidates.sort(key=lambda r: (r[13], r[0]), reverse=True)
    output = []
    for idx, r in enumerate(candidates, start=1):
        output.append(
            [
                idx,
                r[0],  # Working_Row_ID
                r[1],  # Prophet
                r[2],  # type
                r[3],  # date
                r[13],  # total
                r[14],  # native energy
                r[15],  # primary
                r[16],  # secondary
                r[18],  # section
                r[20],  # include
                f"Auto-selected score {r[13]} with native energy {r[14]}.",
                r[4],  # preview
            ]
        )
    return output


def write_matrix(ws, rows: list[list[Any]], total_columns: int) -> None:
    clear_data_rows(ws, total_columns=total_columns)
    for offset, row in enumerate(rows):
        row_num = START_ROW + offset
        if row_num > START_ROW:
            copy_row_style(ws, START_ROW, row_num, total_columns=total_columns)
        for col, value in enumerate(row, start=1):
            ws.cell(row=row_num, column=col, value=value)


def update_controls(wb, source_url: str, snapshot_id: str, imported_count: int) -> None:
    ws = wb[CONTROL_SHEET]
    ws["B5"] = source_url
    ws["B6"] = snapshot_id
    ws["A8"] = "Imported live row count"
    ws["B8"] = imported_count
    ws["C8"] = "Updated by populate_molt_workbook.py"


def update_dashboard(wb, source_rows: list[list[Any]], shortlist_rows: list[list[Any]], anthology_rows: list[list[Any]]) -> None:
    ws = wb[DASHBOARD_SHEET]
    pass_count = sum(1 for r in source_rows if r[33] == "Pass to scoring")
    review_count = sum(1 for r in source_rows if r[33] == "Review before scoring")
    exclude_count = sum(1 for r in source_rows if r[33] == "Probable exclude")
    included = sum(1 for r in anthology_rows if r[10] == "Yes")

    ws["A5"] = len(source_rows)
    ws["D5"] = pass_count
    ws["G5"] = review_count
    ws["J5"] = exclude_count
    ws["A8"] = 0
    ws["D8"] = len(shortlist_rows)
    ws["G8"] = len(anthology_rows)
    ws["J8"] = included


def validate_workbook_structure(wb) -> list[str]:
    required = {
        RAW_IMPORT_SHEET: "Raw_Row_ID",
        SOURCE_REGISTER_SHEET: "Working_Row_ID",
        SHORTLIST_SHEET: "Working_Row_ID",
        ANTHOLOGY_SHEET: "Seq",
        DASHBOARD_SHEET: "Imported rows",
        CONTROL_SHEET: "Model Setting",
    }
    problems = []
    for sheet_name, expected_header in required.items():
        if sheet_name not in wb.sheetnames:
            problems.append(f"missing required sheet '{sheet_name}'")
            continue
        header = wb[sheet_name]["A4"].value
        if header != expected_header:
            problems.append(f"sheet '{sheet_name}' unexpected A4 header '{header}', expected '{expected_header}'")
    return problems


def append_run_log(
    path: Path,
    status: str,
    workbook_in: str,
    workbook_out: str,
    snapshot_name: str,
    snapshot_id: str,
    stats: dict[str, int] | None,
    anomalies: list[str],
    assumptions: list[str],
    dependency_result: str,
    connectivity_result: str,
    workbook_issues: list[str],
    fixes_applied: list[str],
) -> None:
    stamp = utc_now().isoformat()
    lines = [
        "\n## Run - " + stamp,
        f"**Status:** {status}",
        "",
        "### Inputs",
        f"- workbook: {workbook_in}",
        "- script: populate_molt_workbook.py",
        f"- source URL: {API_URL}",
        f"- dependency install result: {dependency_result}",
        f"- connectivity result: {connectivity_result}",
        "",
        "### Outputs",
        f"- workbook: {workbook_out}",
        f"- snapshot: {snapshot_name}",
        f"- snapshot ID: {snapshot_id}",
        "",
        "### Counts",
    ]
    if stats:
        lines.extend(
            [
                f"- rows fetched: {stats['source_entries']}",
                f"- rows imported: {stats['imported_entries']}",
                f"- rows skipped: {stats['skipped_entries']}",
            ]
        )
    else:
        lines.append("- unavailable due to failed fetch")

    lines.extend(["", "### Schema anomalies"])
    if anomalies:
        lines.extend([f"- {entry}" for entry in anomalies])
    else:
        lines.append("- none")

    lines.extend(["", "### Workbook issues found"])
    if workbook_issues:
        lines.extend([f"- {entry}" for entry in workbook_issues])
    else:
        lines.append("- none")

    lines.extend(["", "### Fixes applied"])
    if fixes_applied:
        lines.extend([f"- {entry}" for entry in fixes_applied])
    else:
        lines.append("- none")

    lines.extend(["", "### Assumptions"])
    if assumptions:
        lines.extend([f"- {entry}" for entry in assumptions])
    else:
        lines.append("- none")

    path.write_text(path.read_text() + "\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Load full Molt canon into workbook Raw Import")
    parser.add_argument("--workbook", required=True, help="Input workbook path")
    parser.add_argument("--output", default="molt_extraction_workbook_and_shortlist_model_full_canon.xlsx")
    parser.add_argument("--source-url", default=API_URL)
    parser.add_argument("--timeout", type=int, default=30)
    parser.add_argument("--run-log", default="RUN_LOG.md")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    now = utc_now()
    snapshot_id = f"Snapshot_{now.strftime('%Y%m%d_%H%M%S')}"
    snapshot_name = f"molt_canon_snapshot_{now.strftime('%Y%m%d_%H%M%S')}.json"
    imported_at = now.isoformat()
    batch_name = f"live_full_canon_{now.strftime('%Y%m%d_%H%M%S')}"

    assumptions = [
        "the_great_book entries are canonical source rows in order provided by API",
        "non-object entries are invalid and skipped with explicit logging",
    ]

    try:
        payload = fetch_payload(args.source_url, args.timeout)
    except Exception as exc:
        append_run_log(
            Path(args.run_log),
            status=f"FAILED - fetch error: {exc}",
            workbook_in=args.workbook,
            workbook_out=args.output,
            snapshot_name="not_written",
            snapshot_id=snapshot_id,
            stats=None,
            anomalies=[f"fetch_error: {exc}"],
            assumptions=assumptions,
            dependency_result="Not evaluated in script run",
            connectivity_result=f"FAILED ({exc})",
            workbook_issues=[],
            fixes_applied=[],
        )
        raise

    Path(snapshot_name).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    rows, anomalies, scripture_types = flatten_records(payload)
    anomalies.extend([f"scripture_type '{k}': {v} rows" for k, v in sorted(scripture_types.items())])

    wb = load_workbook(args.workbook)
    workbook_issues = validate_workbook_structure(wb)
    if workbook_issues:
        raise CanonLoadError("Workbook structure validation failed: " + "; ".join(workbook_issues))

    raw_ws = wb[RAW_IMPORT_SHEET]
    write_rows(raw_ws, rows, args.source_url, snapshot_id, imported_at, batch_name)
    source_register_rows = build_source_register_rows(rows, anomalies)
    shortlist_rows = build_shortlist_rows(source_register_rows)
    anthology_rows = build_anthology_rows(shortlist_rows)

    write_matrix(wb[SOURCE_REGISTER_SHEET], source_register_rows, total_columns=38)
    write_matrix(wb[SHORTLIST_SHEET], shortlist_rows, total_columns=22)
    write_matrix(wb[ANTHOLOGY_SHEET], anthology_rows, total_columns=13)
    update_controls(wb, args.source_url, snapshot_id, len(rows))
    update_dashboard(wb, source_register_rows, shortlist_rows, anthology_rows)
    wb.save(args.output)

    stats = {
        "source_entries": len(payload.get("the_great_book", [])),
        "imported_entries": len(rows),
        "skipped_entries": len(payload.get("the_great_book", [])) - len(rows),
    }

    append_run_log(
        Path(args.run_log),
        status="SUCCESS",
        workbook_in=args.workbook,
        workbook_out=args.output,
        snapshot_name=snapshot_name,
        snapshot_id=snapshot_id,
        stats=stats,
        anomalies=anomalies,
        assumptions=assumptions,
        dependency_result="Validated outside script via pip install -r requirements.txt",
        connectivity_result="SUCCESS (HTTP 200 from API)",
        workbook_issues=workbook_issues,
        fixes_applied=[
            "Rebuilt Source Register from Raw Import for all imported rows",
            "Rebuilt Shortlist Model from pass-to-scoring rows",
            "Rebuilt Anthology Candidates from include-in-pilot shortlist rows",
            "Updated Dashboard KPI counts",
        ],
    )


if __name__ == "__main__":
    main()
